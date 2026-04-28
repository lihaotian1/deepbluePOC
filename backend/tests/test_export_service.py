from io import BytesIO

from openpyxl import load_workbook

from app.schemas import CompareRow, OtherRequirementRow
from app.services.export_service import build_export_workbook


EXPECTED_HEADERS = (
    "序号",
    "章节标题",
    "询价文件原文段落或句子",
    "知识库标准化配套条目的原文",
    "差异结论",
    "详细差异说明",
    "分类",
    "审核意见",
    "审核状态",
)


def _load_first_sheet_rows(blob: bytes) -> list[tuple[object, ...]]:
    workbook = load_workbook(BytesIO(blob))
    sheet = workbook.worksheets[0]
    return list(sheet.iter_rows(values_only=True))


def _load_sheet_rows(blob: bytes, title: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(BytesIO(blob))
    return list(workbook[title].iter_rows(values_only=True))


def test_export_workbook_writes_expected_headers_and_compare_rows() -> None:
    rows = [
        CompareRow(
            row_id="row-1",
            chapter_title="6 DOCUMENTATION",
            source_excerpt="Vendor shall provide the appendices in Russian and English.",
            kb_entry_id="General Specification-12",
            kb_entry_text="产品资料仅提供中英文版本。",
            difference_summary_brief="文件语言要求超出我方标准范围。",
            difference_summary="存在冲突：甲方要求附录提供俄语和英语版本，而我方标准仅支持中英文，需要与甲方澄清。",
            type_code="P",
            review_comment="已提醒销售澄清语言范围。",
            review_status="已审",
        )
    ]

    other_requirements = [
        OtherRequirementRow(
            row_id="other-1",
            chapter_title="7 PARAMETERS",
            source_excerpt="Pump flow shall be 120 m3/h.",
            summary="询价文件要求泵流量达到 120 m3/h。",
            source_order=3,
        )
    ]

    workbook_bytes = build_export_workbook(
        rows=rows,
        other_requirements=other_requirements,
        title="标准化配套结果",
    )
    exported_rows = _load_first_sheet_rows(workbook_bytes)
    other_rows = _load_sheet_rows(workbook_bytes, "其他要求")

    assert exported_rows[0] == EXPECTED_HEADERS
    assert exported_rows[1] == (
        1,
        "6 DOCUMENTATION",
        "Vendor shall provide the appendices in Russian and English.",
        "产品资料仅提供中英文版本。",
        "文件语言要求超出我方标准范围。",
        "存在冲突：甲方要求附录提供俄语和英语版本，而我方标准仅支持中英文，需要与甲方澄清。",
        "P",
        "已提醒销售澄清语言范围。",
        "已审",
    )
    assert len(exported_rows) == 2
    assert other_rows[0] == ("序号", "询价文件最小原文", "AI一句话总结")
    assert other_rows[1] == (1, "Pump flow shall be 120 m3/h.", "询价文件要求泵流量达到 120 m3/h。")


def test_export_workbook_does_not_append_other_rows_for_unmatched_content() -> None:
    rows = []

    workbook_bytes = build_export_workbook(rows=rows, other_requirements=[], title="标准化配套结果")
    exported_rows = _load_first_sheet_rows(workbook_bytes)
    other_rows = _load_sheet_rows(workbook_bytes, "其他要求")

    assert exported_rows[0] == EXPECTED_HEADERS
    assert len(exported_rows) == 1
    assert other_rows[0] == ("序号", "询价文件最小原文", "AI一句话总结")
    assert len(other_rows) == 1
